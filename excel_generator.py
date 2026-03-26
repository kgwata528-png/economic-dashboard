"""
Excel生成モジュール
シート構成:
  1. 時系列データ   - 選択指標の価格データ（日付×指標）
  2. 変化率(%)      - 前期比の変化率
  3. 統計サマリー   - 基本統計量
  4. 相関行列       - 指標間の相関
  5. CPI            - 消費者物価指数（月次・前年比）
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ====================================================
# カラーパレット
# ====================================================
C = {
    "header_dark":  "1A237E",  # 濃紺
    "header_mid":   "1976D2",  # 青
    "header_green": "1B5E20",  # 緑
    "header_red":   "B71C1C",  # 赤
    "header_orange":"E65100",  # オレンジ
    "header_purple":"4A148C",  # 紫
    "header_indigo":"283593",  # インディゴ（短観）
    "white":        "FFFFFF",
    "light_blue":   "E3F2FD",
    "light_gray":   "F5F5F5",
    "gray":         "EEEEEE",
}

thin_side  = Side(style="thin",   color="CCCCCC")
thick_side = Side(style="medium", color="9E9E9E")
thin_border  = Border(left=thin_side,  right=thin_side,  top=thin_side,  bottom=thin_side)
thick_border = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)

INTERVAL_LABEL = {"1d": "日足", "1wk": "週足", "1mo": "月足"}
PERIOD_LABEL   = {
    "3mo": "3ヶ月", "6mo": "6ヶ月", "1y": "1年",
    "3y": "3年",   "5y": "5年",   "10y": "10年",
}

# ====================================================
# スタイルヘルパー
# ====================================================
def _font(bold=False, size=10, color="000000", name="Meiryo UI"):
    return Font(name=name, bold=bold, size=size, color=color)

def _fill(color: str):
    return PatternFill("solid", start_color=color)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _write_header(cell, value, bg=C["header_dark"], fg=C["white"], size=10, bold=True, wrap=True):
    cell.value     = value
    cell.font      = _font(bold=bold, size=size, color=fg)
    cell.fill      = _fill(bg)
    cell.alignment = _align(wrap=wrap)
    cell.border    = thin_border

def _write_data(cell, value, bg=C["white"], h_align="center", bold=False,
                num_fmt=None, color="000000"):
    cell.value     = value
    cell.font      = _font(bold=bold, color=color)
    cell.fill      = _fill(bg)
    cell.alignment = _align(h=h_align)
    cell.border    = thin_border
    if num_fmt:
        cell.number_format = num_fmt

def _title_row(ws, row, col_span, text, bg=C["header_dark"], size=13):
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=col_span
    )
    c = ws.cell(row=row, column=1)
    c.value     = text
    c.font      = _font(bold=True, size=size, color=C["white"])
    c.fill      = _fill(bg)
    c.alignment = _align(h="left")
    c.border    = thin_border
    ws.row_dimensions[row].height = 26

def _note_row(ws, row, col_span, text):
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=col_span
    )
    c = ws.cell(row=row, column=1)
    c.value     = text
    c.font      = Font(name="Meiryo UI", size=8, italic=True, color="757575")
    c.fill      = _fill("FAFAFA")
    c.alignment = _align(h="left")
    c.border    = thin_border
    ws.row_dimensions[row].height = 15

# ====================================================
# シート1: 時系列データ
# ====================================================
def _sheet_timeseries(wb: Workbook, df: pd.DataFrame, period: str, interval: str,
                      cpi_df: pd.DataFrame | None = None,
                      tankan_df: pd.DataFrame | None = None):
    ws = wb.active
    ws.title = "時系列データ"
    ws.sheet_properties.tabColor = C["header_mid"]

    # CPIデータを市場データの日付に合わせてforward-fill
    cpi_aligned = None
    cpi_cols = []
    if cpi_df is not None and not cpi_df.empty:
        cpi_df = cpi_df.copy()
        cpi_df.index = pd.to_datetime(cpi_df.index).tz_localize(None)
        df.index = pd.to_datetime(df.index).tz_localize(None) if df.index.tz is not None else pd.to_datetime(df.index)
        combined_idx = cpi_df.index.union(df.index).sort_values()
        cpi_aligned = cpi_df.reindex(combined_idx).ffill().reindex(df.index)
        cpi_cols = cpi_aligned.columns.tolist()

    # 短観データを市場データの日付に合わせてforward-fill（四半期データを日次/週次/月次に展開）
    tankan_aligned = None
    tankan_cols = []
    if tankan_df is not None and not tankan_df.empty:
        _tk = tankan_df.copy()
        _tk.index = pd.to_datetime(_tk.index).tz_localize(None)
        df.index = pd.to_datetime(df.index).tz_localize(None) if df.index.tz is not None else pd.to_datetime(df.index)
        _combined_idx = _tk.index.union(df.index).sort_values()
        tankan_aligned = _tk.reindex(_combined_idx).ffill().reindex(df.index)
        tankan_cols = tankan_aligned.columns.tolist()

    # 欠損値を前の値で補完（土日・祝日・データなし日を前営業日の値で埋める）
    df = df.ffill()

    cols = df.columns.tolist()
    n_cols = len(cols) + len(cpi_cols) + len(tankan_cols) + 1  # +1 for date

    period_str   = PERIOD_LABEL.get(period, period)
    interval_str = INTERVAL_LABEL.get(interval, interval)

    if not df.empty:
        date_range = f"{df.index[0].strftime('%Y/%m/%d')} 〜 {df.index[-1].strftime('%Y/%m/%d')}"
    else:
        date_range = "データなし"

    _title_row(ws, 1, n_cols,
               f"経済指標 時系列データ｜{period_str}・{interval_str}｜{date_range}",
               bg=C["header_dark"])
    _note_row(ws, 2, n_cols,
              "出典: Yahoo Finance（yfinance） ／ FRED（マクロ指標は月次・前年同月比） ／ 土日・祝日は市場休場のため欠損あり")

    # ヘッダー行
    _write_header(ws.cell(row=3, column=1), "日付", bg=C["header_dark"])
    for ci, col in enumerate(cols, 2):
        _write_header(ws.cell(row=3, column=ci), col, bg=C["header_mid"])
    # CPIヘッダー（緑色で区別）
    cpi_start_col = len(cols) + 2
    for ci, col in enumerate(cpi_cols, cpi_start_col):
        _write_header(ws.cell(row=3, column=ci), col, bg=C["header_green"])
    # 短観ヘッダー（インディゴ色で区別）
    tankan_start_col = len(cols) + len(cpi_cols) + 2
    for ci, col in enumerate(tankan_cols, tankan_start_col):
        _write_header(ws.cell(row=3, column=ci), col, bg=C["header_indigo"])
    ws.row_dimensions[3].height = 24

    # データ行
    for ri, (idx, row) in enumerate(df.iterrows(), 4):
        bg = C["light_gray"] if ri % 2 == 0 else C["white"]
        dt_str = idx.strftime("%Y/%m/%d") if hasattr(idx, "strftime") else str(idx)
        _write_data(ws.cell(row=ri, column=1), dt_str, bg=bg)

        for ci, col in enumerate(cols, 2):
            val = row[col]
            if pd.isna(val):
                _write_data(ws.cell(row=ri, column=ci), "―", bg=bg)
            else:
                _write_data(ws.cell(row=ri, column=ci), round(float(val), 4),
                            bg=bg, num_fmt="#,##0.0000")

        # CPIデータ列
        if cpi_aligned is not None:
            for ci, col in enumerate(cpi_cols, cpi_start_col):
                val = cpi_aligned.loc[idx, col] if idx in cpi_aligned.index else None
                if val is None or pd.isna(val):
                    _write_data(ws.cell(row=ri, column=ci), "―", bg=bg)
                else:
                    _write_data(ws.cell(row=ri, column=ci), round(float(val), 2),
                                bg=bg, num_fmt="#,##0.00")

        # 短観データ列
        if tankan_aligned is not None:
            for ci, col in enumerate(tankan_cols, tankan_start_col):
                val = tankan_aligned.loc[idx, col] if idx in tankan_aligned.index else None
                if val is None or (hasattr(val, "__float__") and pd.isna(val)):
                    _write_data(ws.cell(row=ri, column=ci), "―", bg=bg)
                else:
                    _write_data(ws.cell(row=ri, column=ci), round(float(val), 1),
                                bg=bg, num_fmt="#,##0.0")
        ws.row_dimensions[ri].height = 17

    # 列幅
    ws.column_dimensions["A"].width = 13
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = ws.cell(row=3, column=1).coordinate + ":" + \
                         ws.cell(row=3, column=n_cols).coordinate


# ====================================================
# シート2: 変化率(%)
# ====================================================
def _sheet_returns(wb: Workbook, df: pd.DataFrame, interval: str):
    ws = wb.create_sheet("変化率(%)")
    ws.sheet_properties.tabColor = C["header_green"]

    ret = df.pct_change() * 100
    cols = ret.columns.tolist()
    n_cols = len(cols) + 1
    interval_str = INTERVAL_LABEL.get(interval, interval)

    _title_row(ws, 1, n_cols,
               f"経済指標 前期比変化率（%）｜{interval_str}",
               bg=C["header_green"])
    _note_row(ws, 2, n_cols, "各期間の前期比変化率 (%) = (当期値 - 前期値) / 前期値 × 100")

    _write_header(ws.cell(row=3, column=1), "日付", bg=C["header_dark"])
    for ci, col in enumerate(cols, 2):
        _write_header(ws.cell(row=3, column=ci), col, bg=C["header_green"])
    ws.row_dimensions[3].height = 24

    for ri, (idx, row) in enumerate(ret.iterrows(), 4):
        bg = C["light_gray"] if ri % 2 == 0 else C["white"]
        dt_str = idx.strftime("%Y/%m/%d") if hasattr(idx, "strftime") else str(idx)
        _write_data(ws.cell(row=ri, column=1), dt_str, bg=bg)

        for ci, col in enumerate(cols, 2):
            val = row[col]
            if pd.isna(val):
                _write_data(ws.cell(row=ri, column=ci), "―", bg=bg)
            else:
                fval = round(float(val), 4)
                color = "1B5E20" if fval > 0 else ("B71C1C" if fval < 0 else "000000")
                _write_data(ws.cell(row=ri, column=ci), fval,
                            bg=bg, num_fmt='+0.0000%;-0.0000%;0.0000%',
                            color=color)
                ws.cell(row=ri, column=ci).value = fval / 100  # Excelのパーセント書式に合わせる

        ws.row_dimensions[ri].height = 17

    ws.column_dimensions["A"].width = 13
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    ws.freeze_panes = "A4"

    # 条件付き書式（カラースケール）
    if not ret.empty and len(ret) > 1:
        for ci in range(2, n_cols + 1):
            col_letter = get_column_letter(ci)
            data_range = f"{col_letter}4:{col_letter}{3 + len(ret)}"
            ws.conditional_formatting.add(
                data_range,
                ColorScaleRule(
                    start_type="min",   start_color="FF4444",
                    mid_type="num",     mid_value=0, mid_color="FFFFFF",
                    end_type="max",     end_color="44BB44",
                )
            )


# ====================================================
# シート3: 統計サマリー
# ====================================================
def _sheet_stats(wb: Workbook, df: pd.DataFrame, period: str, interval: str):
    ws = wb.create_sheet("統計サマリー")
    ws.sheet_properties.tabColor = C["header_orange"]

    period_str   = PERIOD_LABEL.get(period, period)
    interval_str = INTERVAL_LABEL.get(interval, interval)
    cols = df.columns.tolist()

    _title_row(ws, 1, len(cols) + 1,
               f"統計サマリー｜{period_str}・{interval_str}",
               bg=C["header_orange"])
    _note_row(ws, 2, len(cols) + 1,
              "各指標の基本統計量。変化率は期間全体の始値→終値の変化率。")

    stat_labels = [
        "データ数", "開始日", "終了日",
        "最新値", "始値", "最高値", "最安値",
        "平均値", "中央値", "標準偏差",
        "期間変化率(%)", "最大上昇率(%)", "最大下落率(%)",
    ]

    _write_header(ws.cell(row=3, column=1), "統計項目", bg=C["header_dark"])
    for ci, col in enumerate(cols, 2):
        _write_header(ws.cell(row=3, column=ci), col, bg=C["header_orange"])
    ws.row_dimensions[3].height = 24

    for ri, label in enumerate(stat_labels, 4):
        bg = C["light_gray"] if ri % 2 == 0 else C["white"]
        _write_data(ws.cell(row=ri, column=1), label, bg=bg, h_align="left", bold=True)

        for ci, col in enumerate(cols, 2):
            s = df[col].dropna()
            if s.empty:
                _write_data(ws.cell(row=ri, column=ci), "データなし", bg=bg)
                continue

            pct = (s.pct_change() * 100).dropna()

            val_map = {
                "データ数":       (f"{len(s)} 本",  None),
                "開始日":         (s.index[0].strftime("%Y/%m/%d") if hasattr(s.index[0], "strftime") else str(s.index[0]), None),
                "終了日":         (s.index[-1].strftime("%Y/%m/%d") if hasattr(s.index[-1], "strftime") else str(s.index[-1]), None),
                "最新値":         (round(float(s.iloc[-1]), 4), "#,##0.0000"),
                "始値":           (round(float(s.iloc[0]),  4), "#,##0.0000"),
                "最高値":         (round(float(s.max()),    4), "#,##0.0000"),
                "最安値":         (round(float(s.min()),    4), "#,##0.0000"),
                "平均値":         (round(float(s.mean()),   4), "#,##0.0000"),
                "中央値":         (round(float(s.median()), 4), "#,##0.0000"),
                "標準偏差":       (round(float(s.std()),    4), "#,##0.0000"),
                "期間変化率(%)":  (round((float(s.iloc[-1]) / float(s.iloc[0]) - 1) * 100, 2) if s.iloc[0] != 0 else 0, "+0.00%;-0.00%"),
                "最大上昇率(%)":  (round(float(pct.max()), 2) if not pct.empty else 0, "+0.00%;-0.00%"),
                "最大下落率(%)":  (round(float(pct.min()), 2) if not pct.empty else 0, "+0.00%;-0.00%"),
            }

            v, fmt = val_map[label]
            if fmt and isinstance(v, float):
                # パーセント系は小数に変換
                if "%" in fmt:
                    _write_data(ws.cell(row=ri, column=ci), v / 100, bg=bg, num_fmt=fmt)
                else:
                    _write_data(ws.cell(row=ri, column=ci), v, bg=bg, num_fmt=fmt)
            else:
                _write_data(ws.cell(row=ri, column=ci), v, bg=bg)

        ws.row_dimensions[ri].height = 20

    ws.column_dimensions["A"].width = 18
    for ci in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 18


# ====================================================
# シート4: 相関行列
# ====================================================
def _sheet_correlation(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("相関行列")
    ws.sheet_properties.tabColor = C["header_purple"]

    cols = df.columns.tolist()
    corr = df.corr()

    _title_row(ws, 1, len(cols) + 1, "指標間の相関行列", bg=C["header_purple"])
    _note_row(ws, 2, len(cols) + 1,
              "ピアソン相関係数 (-1〜+1)。1に近いほど強い正の相関、-1に近いほど強い負の相関。")

    _write_header(ws.cell(row=3, column=1), "", bg=C["header_dark"])
    for ci, col in enumerate(cols, 2):
        _write_header(ws.cell(row=3, column=ci), col, bg=C["header_purple"])
        _write_header(ws.cell(row=ci + 2, column=1), col, bg=C["header_purple"])
    ws.row_dimensions[3].height = 24

    for ri, row_col in enumerate(cols, 4):
        ws.row_dimensions[ri].height = 20
        for ci, col_col in enumerate(cols, 2):
            val = corr.loc[row_col, col_col] if (row_col in corr.index and col_col in corr.columns) else None
            if val is None or pd.isna(val):
                _write_data(ws.cell(row=ri, column=ci), "―")
            elif row_col == col_col:
                _write_data(ws.cell(row=ri, column=ci), 1.0,
                            bg=C["light_blue"], num_fmt="0.00", bold=True)
            else:
                fval = round(float(val), 2)
                _write_data(ws.cell(row=ri, column=ci), fval, num_fmt="0.00")

    ws.column_dimensions["A"].width = 18
    for ci in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    # カラースケール
    if len(cols) > 0:
        end_col = get_column_letter(len(cols) + 1)
        end_row = 3 + len(cols)
        data_range = f"B4:{end_col}{end_row}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="num",  start_value=-1, start_color="FF6666",
                mid_type="num",    mid_value=0,    mid_color="FFFFFF",
                end_type="num",    end_value=1,    end_color="4CAF50",
            )
        )


# ====================================================
# シート5: CPI
# ====================================================
def _sheet_cpi(wb: Workbook, cpi_df: pd.DataFrame):
    ws = wb.create_sheet("マクロ経済指標（FRED）")
    ws.sheet_properties.tabColor = C["header_red"]

    cols = cpi_df.columns.tolist()
    n_cols = len(cols) + 1

    _title_row(ws, 1, n_cols,
               "マクロ経済指標｜CPI・失業率・NFP・GDP（FRED）",
               bg=C["header_red"])
    _note_row(ws, 2, n_cols,
              "出典: FRED (Federal Reserve Bank of St. Louis)／月次・四半期データ　※CPI/GDP は前年同月比(%)、NFP は前月差(千人)")

    _write_header(ws.cell(row=3, column=1), "年月", bg=C["header_dark"])
    for ci, col in enumerate(cols, 2):
        _write_header(ws.cell(row=3, column=ci), col, bg=C["header_red"])
    ws.row_dimensions[3].height = 24

    for ri, (idx, row) in enumerate(cpi_df.iterrows(), 4):
        bg = C["light_gray"] if ri % 2 == 0 else C["white"]
        dt_str = idx.strftime("%Y/%m") if hasattr(idx, "strftime") else str(idx)
        _write_data(ws.cell(row=ri, column=1), dt_str, bg=bg)
        for ci, col in enumerate(cols, 2):
            val = row[col]
            if pd.isna(val):
                _write_data(ws.cell(row=ri, column=ci), "―", bg=bg)
            else:
                fval = round(float(val), 2)
                color = "B71C1C" if fval > 2 else ("1B5E20" if fval < 0 else "000000")
                _write_data(ws.cell(row=ri, column=ci), fval / 100,
                            bg=bg, num_fmt="+0.00%;-0.00%", color=color)
        ws.row_dimensions[ri].height = 18

    ws.column_dimensions["A"].width = 12
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    ws.freeze_panes = "A4"

    # カラースケール
    for ci in range(2, n_cols + 1):
        col_letter = get_column_letter(ci)
        data_range = f"{col_letter}4:{col_letter}{3 + len(cpi_df)}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min",   start_color="4CAF50",
                mid_type="num",     mid_value=0.02, mid_color="FFFFFF",
                end_type="max",     end_color="F44336",
            )
        )


def _sheet_tankan(wb: Workbook, tankan_df: pd.DataFrame):
    """日銀短観シート"""
    ws = wb.create_sheet("日銀短観（業況判断DI）")
    ws.sheet_view.showGridLines = False
    cols = tankan_df.columns.tolist()
    n_cols = len(cols) + 1

    _title_row(ws, 1, n_cols,
               "日銀短観（業況判断DI）",
               bg=C["header_dark"])
    _note_row(ws, 2, n_cols,
              "出典: 日本銀行 短観（全国企業短期経済観測調査）／四半期データ"
              "　※DI＝「良い」－「悪い」（%ポイント）、プラスは景況改善を示す")

    _write_header(ws.cell(row=3, column=1), "四半期", bg=C["header_dark"])
    for ci, col in enumerate(cols, 2):
        _write_header(ws.cell(row=3, column=ci), col, bg=C["header_dark"])
    ws.row_dimensions[3].height = 28

    for ri, (idx, row) in enumerate(tankan_df.iterrows(), 4):
        bg = C["light_gray"] if ri % 2 == 0 else C["white"]
        q = (idx.month - 1) // 3 + 1
        dt_str = f"{idx.year}/Q{q}"
        _write_data(ws.cell(row=ri, column=1), dt_str, bg=bg)
        for ci, col in enumerate(cols, 2):
            val = row[col]
            if pd.isna(val):
                _write_data(ws.cell(row=ri, column=ci), "\u2015", bg=bg)
            else:
                fval = round(float(val), 1)
                color = "2E7D32" if fval > 0 else ("C62828" if fval < 0 else "000000")
                _write_data(ws.cell(row=ri, column=ci), fval,
                            bg=bg, num_fmt="+0.0;-0.0;0.0", color=color)
        ws.row_dimensions[ri].height = 18

    ws.column_dimensions["A"].width = 12
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 24



# ====================================================
# メイン生成関数
# ====================================================
def generate_excel(
    market_df: pd.DataFrame,
    cpi_df: pd.DataFrame | None,
    period: str,
    interval: str,
    tankan_df: pd.DataFrame | None = None,
) -> bytes:
    """Excel ファイルをバイト列で返す"""
    wb = Workbook()

    if market_df.empty:
        ws = wb.active
        ws.title = "エラー"
        ws["A1"] = "データを取得できませんでした。指標を選択してください。"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # 短観データをmarket_dfに結合してcombined_dfを作成（①③④内の相連分析に使用）
    combined_df = market_df.copy()
    if tankan_df is not None and not tankan_df.empty:
        _tk2 = tankan_df.copy()
        _tk2.index = pd.to_datetime(_tk2.index).tz_localize(None)
        _mkt_idx = pd.to_datetime(combined_df.index).tz_localize(None) if combined_df.index.tz is not None else pd.to_datetime(combined_df.index)
        combined_df.index = _mkt_idx
        _cidx = _tk2.index.union(combined_df.index).sort_values()
        _tk2_aligned = _tk2.reindex(_cidx).ffill().reindex(combined_df.index)
        combined_df = pd.concat([combined_df, _tk2_aligned], axis=1)

    _sheet_timeseries(wb, market_df, period, interval, cpi_df, tankan_df=tankan_df)
    _sheet_returns(wb, combined_df, interval)
    _sheet_stats(wb, combined_df, period, interval)

    if len(combined_df.columns) > 1:
        _sheet_correlation(wb, combined_df)

    if cpi_df is not None and not cpi_df.empty:
        _sheet_cpi(wb, cpi_df)

    if tankan_df is not None and not tankan_df.empty:
        _sheet_tankan(wb, tankan_df)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
