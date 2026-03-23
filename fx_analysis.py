"""
為替・米長期金利データ取得 & Excel作成
======================================
【インストール】
    pip install yfinance openpyxl pandas

【実行】
    python fetch_fx_bond.py

【出力】
    fx_bond_data.xlsx（同フォルダに生成）

【取得データ】
    - USD/JPY 1時間足 始値 168本（過去7日）
    - 米10年債利回り 1時間足（取れない場合は日次で代替）
"""

import sys
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 設定
# ============================================================
OUTPUT_FILE = "fx_bond_data.xlsx"
END   = datetime.now(timezone.utc)
START = END - timedelta(days=7)

# ============================================================
# 1. USD/JPY 1時間足（始値168本）
# ============================================================
print("【1】USD/JPY 1時間足を取得中...")
usdjpy_raw = yf.download("USDJPY=X", start=START, end=END,
                          interval="1h", auto_adjust=True, progress=False)

if usdjpy_raw.empty:
    print("  ❌ USD/JPY の取得に失敗しました。ネット接続を確認してください。")
    sys.exit(1)

# MultiIndex 解除
if isinstance(usdjpy_raw.columns, pd.MultiIndex):
    usdjpy_raw.columns = usdjpy_raw.columns.get_level_values(0)

# タイムゾーン除去・始値のみ抽出
usdjpy_raw.index = pd.to_datetime(usdjpy_raw.index).tz_localize(None) \
    if usdjpy_raw.index.tz is None else pd.to_datetime(usdjpy_raw.index).tz_convert(None)

usdjpy = usdjpy_raw[["Open"]].rename(columns={"Open": "USD/JPY 始値"}).round(3)
print(f"  ✅ 取得: {len(usdjpy)} 本  （{usdjpy.index[0]} 〜 {usdjpy.index[-1]}）")
if len(usdjpy) < 168:
    print(f"  ⚠️  168本未満です（土日・祝日は市場が薄く欠損することがあります）")

# ============================================================
# 2. 米10年債利回り（1時間足 → 取れなければ日次）
# ============================================================
print("\n【2】米10年債利回りを取得中...")
us10y = None
interval_used = None

# まず1時間足を試みる
us10y_raw = yf.download("^TNX", start=START, end=END,
                          interval="1h", auto_adjust=True, progress=False)
if isinstance(us10y_raw.columns, pd.MultiIndex):
    us10y_raw.columns = us10y_raw.columns.get_level_values(0)

if not us10y_raw.empty and "Close" in us10y_raw.columns:
    us10y_raw.index = pd.to_datetime(us10y_raw.index).tz_localize(None) \
        if us10y_raw.index.tz is None else pd.to_datetime(us10y_raw.index).tz_convert(None)
    us10y = us10y_raw[["Close"]].rename(columns={"Close": "米10年債利回り(%)"}).round(4)
    interval_used = "1時間足"
    print(f"  ✅ 1時間足で取得: {len(us10y)} 本")
else:
    print("  ⚠️  1時間足が取得できませんでした。日次で再試行...")
    us10y_raw = yf.download("^TNX", start=START - timedelta(days=3), end=END,
                             interval="1d", auto_adjust=True, progress=False)
    if isinstance(us10y_raw.columns, pd.MultiIndex):
        us10y_raw.columns = us10y_raw.columns.get_level_values(0)

    if not us10y_raw.empty and "Close" in us10y_raw.columns:
        us10y_raw.index = pd.to_datetime(us10y_raw.index).tz_localize(None) \
            if us10y_raw.index.tz is None else pd.to_datetime(us10y_raw.index).tz_convert(None)
        us10y = us10y_raw[["Close"]].rename(columns={"Close": "米10年債利回り(%)"}).round(4)
        interval_used = "日次"
        print(f"  ✅ 日次で取得: {len(us10y)} 本")
    else:
        print("  ❌ 米10年債利回りの取得に失敗しました（^TNX が利用不可の可能性）")

# ============================================================
# 3. Excel 作成
# ============================================================
print("\n【3】Excel を作成中...")

# --- スタイル定数 ---
C_DARK   = "1F3864"
C_MID    = "2E75B6"
C_ORANGE = "C65911"
C_GREEN  = "375623"
C_WHITE  = "FFFFFF"
C_LGRAY  = "F2F2F2"

thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, value, bg=C_DARK, fg=C_WHITE, sz=10, bold=True):
    cell.value     = value
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border

def dat(cell, value, bg=C_WHITE, align="center", bold=False, color="000000", fmt=None):
    cell.value     = value
    cell.font      = Font(name="Arial", size=9, bold=bold, color=color)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = border
    if fmt:
        cell.number_format = fmt

wb = Workbook()

# ============================================================
# シート1: メインデータ（USD/JPY 始値 + 米10年債）
# ============================================================
ws1 = wb.active
ws1.title = "為替・金利データ"
ws1.sheet_properties.tabColor = C_MID

period_str = f"{usdjpy.index[0].strftime('%Y/%m/%d')} 〜 {usdjpy.index[-1].strftime('%Y/%m/%d')}"
bond_note  = f"米10年債: {interval_used}データ" if interval_used else "米10年債: 取得不可"

# タイトル
ws1.merge_cells("A1:C1")
c = ws1["A1"]
c.value     = f"USD/JPY 始値（1時間足）・米10年債利回り　{period_str}　　出典: Yahoo Finance"
c.font      = Font(name="Arial", bold=True, size=12, color=C_WHITE)
c.fill      = PatternFill("solid", start_color=C_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

# 注記
ws1.merge_cells("A2:C2")
c2 = ws1["A2"]
c2.value     = f"※ USD/JPY: 1時間足 始値 {len(usdjpy)}本　　{bond_note}　　土日・祝日は市場休場のため欠損あり"
c2.font      = Font(name="Arial", size=8, italic=True, color="666666")
c2.fill      = PatternFill("solid", start_color="EEF2F7")
c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws1.row_dimensions[2].height = 16

# ヘッダー
hdr(ws1["A3"], "日時",            bg=C_DARK)
hdr(ws1["B3"], "USD/JPY 始値（円）", bg=C_MID)
hdr(ws1["C3"], f"米10年債利回り（%）\n{interval_used or '取得不可'}", bg=C_ORANGE)
ws1.row_dimensions[3].height = 30

# --- データ行の組み立て ---
# USD/JPY をベースに、米10年債をマージ（日次の場合は日付ベースでマッチ）
df_main = usdjpy.copy()
df_main.index.name = "日時"

if us10y is not None:
    if interval_used == "1時間足":
        # インデックスを揃えて結合
        df_main = df_main.join(us10y, how="left")
    else:
        # 日次の場合: 日付列を作ってマージ
        us10y_daily = us10y.copy()
        us10y_daily.index = pd.to_datetime(us10y_daily.index).normalize()
        df_main["_date"] = pd.to_datetime(df_main.index).normalize()
        df_main = df_main.merge(us10y_daily, left_on="_date", right_index=True, how="left")
        df_main = df_main.drop(columns=["_date"])
        df_main.index = usdjpy.index  # インデックスを元に戻す

if "米10年債利回り(%)" not in df_main.columns:
    df_main["米10年債利回り(%)"] = None

# データ書き込み
for ri, (idx, row) in enumerate(df_main.iterrows(), 4):
    bg = C_LGRAY if ri % 2 == 0 else C_WHITE
    dt_val = idx.strftime("%Y/%m/%d %H:%M") if hasattr(idx, "strftime") else str(idx)

    dat(ws1.cell(row=ri, column=1), dt_val,                   bg=bg)
    dat(ws1.cell(row=ri, column=2), row["USD/JPY 始値"],      bg=bg, fmt="0.000")
    bond_val = row.get("米10年債利回り(%)") if isinstance(row, pd.Series) else row["米10年債利回り(%)"] if "米10年債利回り(%)" in df_main.columns else None
    dat(ws1.cell(row=ri, column=3), bond_val if pd.notna(bond_val) else "―",
        bg=bg, fmt="0.0000" if pd.notna(bond_val) else None)
    ws1.row_dimensions[ri].height = 17

ws1.column_dimensions["A"].width = 18
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 18
ws1.freeze_panes = "A4"

# ============================================================
# シート2: 統計サマリー
# ============================================================
ws2 = wb.create_sheet("統計サマリー")
ws2.sheet_properties.tabColor = C_GREEN

ws2.merge_cells("A1:D1")
c = ws2["A1"]
c.value     = f"統計サマリー　{period_str}　　出典: Yahoo Finance"
c.font      = Font(name="Arial", bold=True, size=12, color=C_WHITE)
c.fill      = PatternFill("solid", start_color=C_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

# USD/JPY セクション
ws2.merge_cells("A3:D3")
s = ws2["A3"]
s.value     = "■ USD/JPY 始値　統計（1時間足）"
s.font      = Font(name="Arial", bold=True, size=11, color=C_WHITE)
s.fill      = PatternFill("solid", start_color=C_MID)
s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.row_dimensions[3].height = 22

close = df_main["USD/JPY 始値"].dropna()
fx_stats = [
    ("データ本数",  f"{len(close)} 本"),
    ("期間",        period_str),
    ("最高値",      f"{close.max():.3f} 円"),
    ("最安値",      f"{close.min():.3f} 円"),
    ("平均値",      f"{close.mean():.3f} 円"),
    ("標準偏差",    f"{close.std():.4f}"),
    ("変動幅",      f"{close.max() - close.min():.3f} 円"),
    ("期間変化率",  f"{(close.iloc[-1] / close.iloc[0] - 1) * 100:+.2f}%"),
]

hdr(ws2.cell(row=4, column=1), "指標",  bg=C_DARK)
hdr(ws2.cell(row=4, column=2), "値",    bg=C_MID)
ws2.merge_cells("C4:D4")
hdr(ws2.cell(row=4, column=3), "",      bg=C_DARK)
ws2.row_dimensions[4].height = 22

for ri, (label, val) in enumerate(fx_stats, 5):
    bg = C_LGRAY if ri % 2 == 0 else C_WHITE
    dat(ws2.cell(row=ri, column=1), label, bg=bg, align="left", bold=True)
    dat(ws2.cell(row=ri, column=2), val,   bg=bg)
    ws2.merge_cells(f"C{ri}:D{ri}")
    dat(ws2.cell(row=ri, column=3), "", bg=bg)
    ws2.row_dimensions[ri].height = 20

# 米10年債セクション
r_start = 5 + len(fx_stats) + 1
ws2.merge_cells(f"A{r_start}:D{r_start}")
s2 = ws2[f"A{r_start}"]
s2.value     = f"■ 米10年債利回り　統計（{interval_used or '取得不可'}）"
s2.font      = Font(name="Arial", bold=True, size=11, color=C_WHITE)
s2.fill      = PatternFill("solid", start_color=C_ORANGE)
s2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.row_dimensions[r_start].height = 22

hdr(ws2.cell(row=r_start+1, column=1), "指標", bg=C_DARK)
hdr(ws2.cell(row=r_start+1, column=2), "値",   bg=C_ORANGE)
ws2.merge_cells(f"C{r_start+1}:D{r_start+1}")
hdr(ws2.cell(row=r_start+1, column=3), "", bg=C_DARK)
ws2.row_dimensions[r_start+1].height = 22

if us10y is not None:
    bond = df_main["米10年債利回り(%)"].dropna()
    bond_stats = [
        ("データ本数", f"{len(bond)} 本（{interval_used}）"),
        ("最高値",     f"{bond.max():.4f}%"),
        ("最安値",     f"{bond.min():.4f}%"),
        ("平均値",     f"{bond.mean():.4f}%"),
        ("標準偏差",   f"{bond.std():.4f}"),
        ("変化幅(bp)", f"{(bond.max() - bond.min()) * 100:.1f} bp"),
        ("期間変化",   f"{(bond.iloc[-1] - bond.iloc[0]) * 100:+.1f} bp"),
    ]
else:
    bond_stats = [("取得結果", "データ取得不可（^TNX が利用できませんでした）")]

for ri, (label, val) in enumerate(bond_stats, r_start+2):
    bg = C_LGRAY if ri % 2 == 0 else C_WHITE
    dat(ws2.cell(row=ri, column=1), label, bg=bg, align="left", bold=True)
    dat(ws2.cell(row=ri, column=2), val,   bg=bg)
    ws2.merge_cells(f"C{ri}:D{ri}")
    dat(ws2.cell(row=ri, column=3), "", bg=bg)
    ws2.row_dimensions[ri].height = 20

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 10

# ============================================================
# 保存
# ============================================================
wb.save(OUTPUT_FILE)
print(f"\n✅ 完成: {OUTPUT_FILE}")
print(f"   シート①「為替・金利データ」: USD/JPY 始値 {len(usdjpy)}本 + 米10年債")
print(f"   シート②「統計サマリー」: 各種統計量")
print(f"\n   ⬆️  このファイルを Claude にアップロードすると、さらに分析・加工できます。")